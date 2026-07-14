#!/usr/bin/env python3
"""
NCD Listing Enrichment - Browser App (Streamlit)
=================================================
Drag-and-drop an Excel file of listed NCDs; the app enriches each ISIN via
the NSDL India Bond Info public API (parallel fetching), normalises issuer
names, and produces a formatted, downloadable workbook with:

    1. Enriched_Data   - original data + enriched columns + clean issuer name
    2. Top_50_Issuers  - ranked by amount raised (live COUNTIFS/SUMIFS formulas)
    3. Exceptions      - ISINs whose API lookups failed or returned empty data

Setup (one-time):
    pip install streamlit pandas requests openpyxl

Run:
    streamlit run ncd_app.py

Notes:
  * A JSON cache (nsdl_cache.json, saved next to this file) stores every
    successful API response - re-processing the same ISINs is instant.
    Your existing cache file from the CLI version is fully compatible:
    just keep it in the same folder as this app.
  * Fetching runs 6 ISINs in parallel (~40-60s for 200 fresh ISINs).
"""

import io
import json
import re
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import pandas as pd

# --------------------------------------------------------------------------
# Configuration
# --------------------------------------------------------------------------
BASE = "https://www.indiabondinfo.nsdl.com/bds-service/v1/public"
ENDPOINTS = {
    "isin":       BASE + "/isins?isin={isin}",
    "ratings":    BASE + "/bdsinfo/credit-ratings?isin={isin}",
    "instrument": BASE + "/bdsinfo/instruments?isin={isin}",
}
CACHE_FILE = Path(__file__).parent / "nsdl_cache.json"
MAX_WORKERS = 6          # parallel ISINs; keep modest to avoid throttling
MAX_RETRIES = 3
TIMEOUT = 20

# --------------------------------------------------------------------------
# Payload validation
# --------------------------------------------------------------------------
def is_empty_payload(d):
    """True when the API returned a JSON body with no usable data
    (every top-level value is null/empty) - treated as a failed lookup."""
    if not isinstance(d, dict) or not d:
        return True
    return all(v in (None, "", [], {}) or (isinstance(v, str) and not v.strip())
               for v in d.values())

# --------------------------------------------------------------------------
# Fetch layer: parallel, cached, thread-safe
# --------------------------------------------------------------------------
class Fetcher:
    def __init__(self):
        import requests
        self._requests = requests
        self._local = threading.local()          # one Session per thread
        self._lock = threading.Lock()
        self.cache = {}
        if CACHE_FILE.exists():
            try:
                raw = json.loads(CACHE_FILE.read_text())
                # Scrub previously-cached empty payloads so they get retried
                self.cache = {k: v for k, v in raw.items() if not is_empty_payload(v)}
            except Exception:
                self.cache = {}
        self._dirty = 0

    def _session(self):
        if not hasattr(self._local, "s"):
            s = self._requests.Session()
            s.headers.update({
                "User-Agent": "Mozilla/5.0 (compatible; NCD-research-script)",
                "Accept": "application/json",
            })
            self._local.s = s
        return self._local.s

    def save_cache(self):
        with self._lock:
            CACHE_FILE.write_text(json.dumps(self.cache))
            self._dirty = 0

    def get(self, kind, isin):
        """Parsed JSON for (kind, isin), or None on failure/empty payload."""
        key = f"{kind}:{isin}"
        with self._lock:
            if key in self.cache:
                return self.cache[key]

        url = ENDPOINTS[kind].format(isin=isin)
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                r = self._session().get(url, timeout=TIMEOUT)
                if r.status_code == 200:
                    data = r.json()
                    if is_empty_payload(data):
                        return None              # 200 but no usable data
                    with self._lock:
                        self.cache[key] = data
                        self._dirty += 1
                        if self._dirty >= 25:    # periodic save
                            CACHE_FILE.write_text(json.dumps(self.cache))
                            self._dirty = 0
                    return data
                elif r.status_code in (429, 502, 503):
                    time.sleep(2 * attempt)
                else:
                    return None
            except Exception:
                time.sleep(2 * attempt)
        return None

    def fetch_isin(self, isin):
        """All three endpoints for one ISIN (runs inside a worker thread)."""
        return isin, {
            "isin":       self.get("isin", isin),
            "ratings":    self.get("ratings", isin),
            "instrument": self.get("instrument", isin),
        }

# --------------------------------------------------------------------------
# Extraction helpers
# --------------------------------------------------------------------------
AGENCY_PATTERNS = [
    (r"\bCRISIL\b",        "CRISIL"),
    (r"\bICRA\b",          "ICRA"),
    (r"\bCARE\b",          "CARE"),
    (r"INDIA\s+RATINGS?\b", "India Ratings"),
    (r"\bFITCH\b",         "India Ratings"),
    (r"\bACUITE\b",        "Acuite"),
    (r"\bBRICKWORK\b",     "Brickwork"),
    (r"\bINFOMERICS\b",    "Infomerics"),
    (r"\bACER\b",          "ACER"),
    (r"\bSMERA\b",         "SMERA"),
]

def short_agency(name):
    up = (name or "").upper()
    for pat, short in AGENCY_PATTERNS:
        if re.search(pat, up):
            return short
    return name.title() if name and name.isupper() else name

def extract_issuer(data):
    if not data:
        return None, None, None
    t = (data.get("issuerTypeOwner") or "").strip() or None
    ind = (data.get("basicIndusrty") or data.get("basicIndustry") or "").strip() or None
    name = (data.get("issuerName") or "").strip() or None
    return t, ind, name

def extract_ratings(data):
    if not data:
        return None
    parts = []
    for r in (data.get("currentRatings") or []):
        agency = short_agency((r.get("creditRatingAgencyName") or "").strip())
        rating = (r.get("currentRating") or "").strip()
        outlook = (r.get("outlook") or "").strip()
        if not rating:
            continue
        s = f"{agency}: {rating}" if agency else rating
        if outlook and outlook != "-":
            s += f" ({outlook})"
        parts.append(s)
    if parts:
        return "; ".join(parts)
    flag = (data.get("instrumentRateFlag") or "").strip()
    return flag or None

def extract_category(data):
    if not data:
        return None
    try:
        cat = data["instrumentsVo"]["instruments"].get("category")
        return cat.strip() if cat and cat.strip() not in ("-", "") else None
    except (KeyError, TypeError, AttributeError):
        return None

# --------------------------------------------------------------------------
# Issuer name normalisation
# --------------------------------------------------------------------------
SUFFIX_MAP = [
    (r"\bLTD\.?\b", "LIMITED"),
    (r"\bPVT\.?\b", "PRIVATE"),
    (r"\bPRIVTE\b", "PRIVATE"),
    (r"\bCORPN\.?\b", "CORPORATION"),
]

def clean_name(name):
    if not isinstance(name, str) or not name.strip():
        return ""
    s = name.upper().strip()
    s = s.replace("&", " AND ")
    s = re.sub(r"[*\u2122\u00ae]", "", s)
    s = re.sub(r"[^\w\s\-\.]", " ", s)
    for pat, rep in SUFFIX_MAP:
        s = re.sub(pat, rep, s)
    s = re.sub(r"\.\s*", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    keep_upper = {"NHPC", "REC", "LIC", "PNB", "IIFL", "HDB", "NIIF", "SMFG",
                  "MAS", "RDC", "ESAF", "NCD", "SDI", "REIT", "INVIT",
                  "NBFC", "L", "T", "A", "K", "R", "AK", "II", "III", "IV"}
    keep_lower = {"AND", "OF", "FOR", "THE", "IN"}
    words = []
    for i, w in enumerate(s.split()):
        if w in keep_upper:
            words.append(w)
        elif w in keep_lower and i > 0:
            words.append(w.lower())
        else:
            words.append(w.capitalize())
    return " ".join(words)

# --------------------------------------------------------------------------
# Core pipeline (UI-independent; progress_cb(done, total) is optional)
# --------------------------------------------------------------------------
def run_pipeline(df, progress_cb=None):
    isin_col = next((c for c in df.columns if str(c).strip().upper() == "ISIN"), None)
    name_col = next((c for c in df.columns if "COMPANY" in str(c).upper()
                     or "ISSUER" in str(c).upper()), None)
    if isin_col is None:
        raise ValueError("No ISIN column found in the uploaded file.")

    df = df.copy()
    df[isin_col] = df[isin_col].astype(str).str.strip().str.upper()
    unique_isins = [i for i in df[isin_col].dropna().unique()
                    if re.match(r"^IN[A-Z0-9]{10}$", i)]

    for col in df.columns:
        if str(col).strip() == "Type of Issue":
            df[col] = (df[col].astype(str).str.strip()
                       .replace({"Reissue": "Re-Issue", "nan": ""}))

    fetcher = Fetcher()
    results = {}
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = [pool.submit(fetcher.fetch_isin, i) for i in unique_isins]
        for n, fut in enumerate(as_completed(futures), 1):
            isin, payloads = fut.result()
            results[isin] = payloads
            if progress_cb:
                progress_cb(n, len(unique_isins))
    fetcher.save_cache()

    enrich, exceptions = {}, []
    for isin in unique_isins:
        p = results[isin]
        t, ind, nsdl_name = extract_issuer(p["isin"])
        enrich[isin] = dict(type_of_issuer=t, industry=ind, nsdl_name=nsdl_name,
                            ratings=extract_ratings(p["ratings"]),
                            category=extract_category(p["instrument"]))
        failed = [lbl for lbl, key in [("issuer details", "isin"),
                                       ("credit ratings", "ratings"),
                                       ("instrument", "instrument")]
                  if p[key] is None]
        if failed:
            src = df.loc[df[isin_col] == isin, name_col].iloc[0] if name_col else ""
            exceptions.append((isin, src, failed))

    df["Type of Issuer"] = df[isin_col].map(lambda i: (enrich.get(i) or {}).get("type_of_issuer"))
    df["Industry of Issuer"] = df[isin_col].map(lambda i: (enrich.get(i) or {}).get("industry"))
    df["Credit Rating(s)"] = df[isin_col].map(lambda i: (enrich.get(i) or {}).get("ratings"))
    df["Category of Instrument"] = df[isin_col].map(lambda i: (enrich.get(i) or {}).get("category"))

    def best_name(row):
        nsdl = (enrich.get(row[isin_col]) or {}).get("nsdl_name")
        return clean_name(nsdl) if nsdl else clean_name(row[name_col]) if name_col else ""
    df["Clean Issuer Name"] = df.apply(best_name, axis=1)

    amt_col = next(c for c in df.columns if "Crores" in str(c))
    top50 = (df.groupby("Clean Issuer Name")[amt_col].sum()
               .sort_values(ascending=False).head(50).index.tolist())
    return df, top50, exceptions

# --------------------------------------------------------------------------
# Excel output (writes to a BytesIO stream for the download button)
# --------------------------------------------------------------------------
def build_workbook_bytes(df, top_n, exceptions):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows

    HDR_FILL = PatternFill("solid", fgColor="1F4E79")
    ALT_FILL = PatternFill("solid", fgColor="DDEBF7")
    HDR_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    BODY_FONT = Font(name="Arial", size=10)
    THIN = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT = Alignment(horizontal="left", vertical="center")

    wb = Workbook()

    # ---------------- Sheet 1: Enriched_Data ----------------
    ws = wb.active
    ws.title = "Enriched_Data"
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    ncols, nrows = df.shape[1], df.shape[0] + 1
    date_cols = {i + 1 for i, c in enumerate(df.columns) if "Date" in str(c)}
    num_fmt_cols = {i + 1 for i, c in enumerate(df.columns) if "Crores" in str(c)}
    for j in range(1, ncols + 1):
        c = ws.cell(row=1, column=j)
        c.font, c.fill, c.border, c.alignment = HDR_FONT, HDR_FILL, BORDER, CENTER
    for i in range(2, nrows + 1):
        for j in range(1, ncols + 1):
            c = ws.cell(row=i, column=j)
            c.font, c.border = BODY_FONT, BORDER
            if j in date_cols and c.value is not None and not isinstance(c.value, str):
                c.number_format = "DD-MM-YYYY"
                c.alignment = CENTER
            elif j in num_fmt_cols:
                c.number_format = "#,##0.00"
            elif isinstance(c.value, (int, float)):
                c.alignment = CENTER
            else:
                c.alignment = LEFT
    for j, col in enumerate(df.columns, 1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df.iloc[:, j - 1].head(200)])
        ws.column_dimensions[get_column_letter(j)].width = min(max(10, max_len + 2), 45)
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{nrows}"

    clean_col_letter = get_column_letter(list(df.columns).index("Clean Issuer Name") + 1)
    amt_name = next(c for c in df.columns if "Crores" in str(c))
    amt_col_letter = get_column_letter(list(df.columns).index(amt_name) + 1)

    # ---------------- Sheet 2: Top_50_Issuers ----------------
    ws2 = wb.create_sheet("Top_50_Issuers")
    ws2["A1"] = "Top 50 Issuers by Amount Raised"
    ws2["A1"].font = Font(name="Arial", size=13, bold=True, color="1F4E79")
    ws2["A3"] = ("Note: All figures are live formulas against the Enriched_Data sheet. "
                 "Counts/amounts use COUNTIFS/SUMIFS on 'Clean Issuer Name' (amounts include "
                 "re-issues). Type and Industry are INDEX/MATCH lookups returning the issuer's "
                 "first occurrence; blank = not populated in NSDL's records.")
    ws2["A3"].font = Font(name="Arial", size=8, italic=True, color="808080")

    headers = ["Rank", "Issuer", "Type of Issuer", "Industry of Issuer",
               "No. of Issuances", "Amount Raised (Rs. Crores)", "% of Total"]
    hdr_row = 5
    for j, h in enumerate(headers, 1):
        c = ws2.cell(row=hdr_row, column=j, value=h)
        c.font, c.fill, c.border, c.alignment = HDR_FONT, HDR_FILL, BORDER, CENTER

    type_col_letter = get_column_letter(list(df.columns).index("Type of Issuer") + 1)
    ind_col_letter = get_column_letter(list(df.columns).index("Industry of Issuer") + 1)
    data_range = f"Enriched_Data!${clean_col_letter}$2:${clean_col_letter}${nrows}"
    amt_range = f"Enriched_Data!${amt_col_letter}$2:${amt_col_letter}${nrows}"
    type_range = f"Enriched_Data!${type_col_letter}$2:${type_col_letter}${nrows}"
    ind_range = f"Enriched_Data!${ind_col_letter}$2:${ind_col_letter}${nrows}"
    first_data, last_data = hdr_row + 1, hdr_row + len(top_n)
    for i, issuer in enumerate(top_n, 0):
        r = first_data + i
        ws2.cell(row=r, column=1, value=i + 1)
        ws2.cell(row=r, column=2, value=issuer)
        ws2.cell(row=r, column=3,
                 value=(f'=IFERROR(IF(INDEX({type_range},MATCH($B{r},{data_range},0))="","",'
                        f'INDEX({type_range},MATCH($B{r},{data_range},0))),"")'))
        ws2.cell(row=r, column=4,
                 value=(f'=IFERROR(IF(INDEX({ind_range},MATCH($B{r},{data_range},0))="","",'
                        f'INDEX({ind_range},MATCH($B{r},{data_range},0))),"")'))
        ws2.cell(row=r, column=5, value=f'=COUNTIFS({data_range},$B{r})')
        ws2.cell(row=r, column=6, value=f'=SUMIFS({amt_range},{data_range},$B{r})')
        ws2.cell(row=r, column=7, value=f'=F{r}/SUM({amt_range})')
        for j in range(1, 8):
            c = ws2.cell(row=r, column=j)
            c.font, c.border = BODY_FONT, BORDER
            if i % 2 == 1:
                c.fill = ALT_FILL
        ws2.cell(row=r, column=1).alignment = CENTER
        ws2.cell(row=r, column=3).alignment = CENTER
        ws2.cell(row=r, column=5).alignment = CENTER
        ws2.cell(row=r, column=6).number_format = "#,##0.00"
        ws2.cell(row=r, column=7).number_format = "0.0%"

    tr = last_data + 1
    ws2.cell(row=tr, column=2, value=f"Total (Top {len(top_n)})")
    ws2.cell(row=tr, column=5, value=f"=SUM(E{first_data}:E{last_data})")
    ws2.cell(row=tr, column=6, value=f"=SUM(F{first_data}:F{last_data})")
    ws2.cell(row=tr, column=7, value=f"=SUM(G{first_data}:G{last_data})")
    for j in range(1, 8):
        c = ws2.cell(row=tr, column=j)
        c.font = Font(name="Arial", size=10, bold=True)
        c.border = Border(top=Side(style="double"), bottom=THIN, left=THIN, right=THIN)
    ws2.cell(row=tr, column=6).number_format = "#,##0.00"
    ws2.cell(row=tr, column=7).number_format = "0.0%"
    for j, w in zip(range(1, 8), [7, 50, 14, 36, 15, 22, 10]):
        ws2.column_dimensions[get_column_letter(j)].width = w
    ws2.freeze_panes = f"A{first_data}"

    # ---------------- Sheet 3: Exceptions ----------------
    ws3 = wb.create_sheet("Exceptions")
    ws3.append(["ISIN", "Company Name (as per source file)", "Failed / Empty Lookups"])
    for j in range(1, 4):
        c = ws3.cell(row=1, column=j)
        c.font, c.fill, c.border, c.alignment = HDR_FONT, HDR_FILL, BORDER, CENTER
    if exceptions:
        for isin, name, failed in exceptions:
            ws3.append([isin, name, ", ".join(failed)])
    else:
        ws3.append(["-", "All ISINs resolved successfully", "-"])
    for row in ws3.iter_rows(min_row=2):
        for c in row:
            c.font, c.border = BODY_FONT, BORDER
    for j, w in zip(range(1, 4), [16, 50, 30]):
        ws3.column_dimensions[get_column_letter(j)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# --------------------------------------------------------------------------
# Streamlit UI
# --------------------------------------------------------------------------
def main():
    import streamlit as st

    st.set_page_config(
        page_title="NCD Data Analysis",
        page_icon="📊",
        layout="wide"
    )

    st.markdown(
        "<h1 style='text-align: center;'>NCD Data Analysis</h1>",
        unsafe_allow_html=True
    )

    uploaded = st.file_uploader("Upload Excel file (must contain an ISIN column)",
                                type=["xlsx", "xls"])
    if uploaded is None:
        cached = len(json.loads(CACHE_FILE.read_text())) if CACHE_FILE.exists() else 0
        if cached:
            st.info(f"Local cache holds {cached} previously fetched API responses - "
                    "already-seen ISINs will resolve instantly.")
        return

    try:
        df_in = pd.read_excel(uploaded)
    except Exception as e:
        st.error(f"Could not read the file: {e}")
        return

    st.write(f"**{len(df_in)} rows** loaded. Preview:")
    st.dataframe(df_in.head(5), use_container_width=True)

    if not st.button("Enrich data", type="primary"):
        return

    prog = st.progress(0.0, text="Fetching from NSDL...")
    def cb(done, total):
        prog.progress(done / total, text=f"Fetching from NSDL... {done}/{total} ISINs")

    t0 = time.time()
    try:
        df_out, top50, exceptions = run_pipeline(df_in, progress_cb=cb)
    except ValueError as e:
        st.error(str(e))
        return
    prog.progress(1.0, text=f"Done in {time.time()-t0:.0f}s")

    c1, c2, c3 = st.columns(3)
    c1.metric("Rows enriched", len(df_out))
    c2.metric("Unique issuers (clean)", df_out["Clean Issuer Name"].nunique())
    c3.metric("Exceptions", len(exceptions))

    if exceptions:
        with st.expander(f"⚠ {len(exceptions)} ISIN(s) with failed/empty lookups"):
            st.dataframe(pd.DataFrame(
                [(i, n, ", ".join(f)) for i, n, f in exceptions],
                columns=["ISIN", "Company Name", "Failed lookups"]),
                use_container_width=True)

    tab1, tab2 = st.tabs(["Enriched data", "Top 50 issuers"])
    with tab1:
        st.dataframe(df_out, use_container_width=True, height=420)
    with tab2:
        amt_col = next(c for c in df_out.columns if "Crores" in str(c))
        top_df = (df_out.groupby("Clean Issuer Name")
                  .agg(Type_of_Issuer=("Type of Issuer", "first"),
                       Industry=("Industry of Issuer", "first"),
                       Issuances=(amt_col, "size"),
                       Amount_Rs_Crores=(amt_col, "sum"))
                  .sort_values("Amount_Rs_Crores", ascending=False).head(50)
                  .reset_index().rename(columns={"Clean Issuer Name": "Issuer"}))
        top_df.index += 1
        st.dataframe(top_df, use_container_width=True, height=420)

    xlsx_bytes = build_workbook_bytes(df_out, top50, exceptions)
    st.download_button(
        "⬇ Download enriched workbook (.xlsx)",
        data=xlsx_bytes,
        file_name="NCD_Enriched_Output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )

if __name__ == "__main__":
    main()
